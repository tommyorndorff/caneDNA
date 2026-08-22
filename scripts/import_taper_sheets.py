#!/usr/bin/env python3
"""Import tapers from the "2019 Bamboo Taper Sheets" workbook (Tom Morgan style).

Source: data/raw/2019_bamboo_taper_sheets.xlsx — one rod per worksheet, each a
"Finished Rod Measurement" sheet with a top-left metadata block (Length, Line
Size, Ferrule, Diameter, Copyright) and a station table split into Tip and Butt
section columns. A guide-station table sits below the taper table and is ignored.

The Tip/Butt header columns vary by sheet, so they are detected per sheet rather
than hard-coded. Each rod is stored losslessly: the raw tip and butt sections are
kept under provenance.extra.sections, and a continuous assembled profile
(tip then butt) is written to `dimensions`/`stations` for plotting.

Output: data/sources/taper_sheets.json
Requires openpyxl. Run build_library.py afterwards to merge into tapers.json.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "raw" / "2019_bamboo_taper_sheets.xlsx"
OUT = ROOT / "data" / "sources" / "taper_sheets.json"

IMPORT_DATE = "2026-08-22"
STATION_RE = re.compile(r"^-?\d+(\.\d+)?\"$")  # e.g. -5"  0"  5"  10"

# Sheets that are templates, not real rods.
SKIP_SHEETS = {"Master"}


def g(row, i):
    return row[i] if i < len(row) else None


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def first_number(s):
    """Extract the first number from a string like "7'", "8.0", "3wt"."""
    if s is None:
        return None
    m = re.search(r"-?\d+(\.\d+)?", str(s))
    return float(m.group()) if m else None


def find_header(rows):
    """Locate (row_idx, tip_col, butt_col) of the Tip/Butt table header."""
    for i, r in enumerate(rows):
        tip_col = butt_col = None
        for j in range(min(len(r), 20)):
            v = g(r, j)
            if isinstance(v, str):
                if v.strip() == "Tip":
                    tip_col = j
                elif v.strip() == "Butt":
                    butt_col = j
        if tip_col is not None and butt_col is not None:
            return i, tip_col, butt_col
    return None


def meta_value(rows, label):
    for r in rows[:8]:
        c0 = g(r, 0)
        if isinstance(c0, str) and c0.strip().lower().startswith(label.lower()):
            return g(r, 1)
    return None


def copyright_text(rows):
    for r in rows[:4]:
        c0 = g(r, 0)
        if isinstance(c0, str) and "opyright" in c0:
            return c0.strip()
    return None


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    hdr = find_header(rows)
    if hdr is None:
        return None
    hrow, tip_col, butt_col = hdr
    tip_dim_col = tip_col + 1  # tip dims sit one column right of the station labels

    tip, butt = [], []
    for r in rows[hrow + 1:]:
        st = g(r, tip_col)
        if not (isinstance(st, str) and STATION_RE.match(st.strip())):
            continue
        pos = float(st.strip().rstrip('"'))
        td, bd = num(g(r, tip_dim_col)), num(g(r, butt_col))
        if td is not None:
            tip.append([pos, td])
        if bd is not None:
            butt.append([pos, bd])

    if not tip or all(d == 0 for _, d in tip):
        return None  # template / empty

    # Assemble a continuous profile: tip section, then butt section shifted to
    # continue after it. Ferrule overlap is not reconciled (raw sections kept).
    inc = 5.0
    dims = [d for _, d in tip]
    stations = [i * inc for i in range(len(tip))]
    if butt:
        offset = stations[-1] + inc
        base = butt[0][0]
        dims += [d for _, d in butt]
        stations += [offset + (p - base) for p, _ in butt]

    length_ft = first_number(meta_value(rows, "Length"))
    line = first_number(meta_value(rows, "Line Size"))
    ferrule = meta_value(rows, "Ferrule")
    diameter = first_number(meta_value(rows, "Diameter"))
    cp = copyright_text(rows)

    provenance = {
        "source": "2019 Bamboo Taper Sheets (Tom Morgan)",
        "author": "Tom W. Morgan",
        "source_url": None,
        "collection": "2019_bamboo_taper_sheets.xlsx",
        "license": cp or "see workbook copyright notice",
        "imported": IMPORT_DATE,
        "sections": {"tip": tip, "butt": butt},
        "const_inferred": True,
        "butt_diameter_in": diameter,
    }

    return {
        "name": ws.title.strip(),
        "type": "Fly-Rod",
        "const_type": "Hex",  # Tom Morgan / PHY rods are hex (inferred)
        "length": round(length_ft * 12, 1) if length_ft else None,
        "line_weight": line,
        "ferrule1_size": str(ferrule).strip() if ferrule else None,
        "station_increment": 5,
        "notes": cp,
        "dimensions": dims,
        "stations": stations,
        "stresses": [],
        "guide_spacings": [],
        "guide_sizes": [],
        "provenance": provenance,
    }


def main():
    if not XLSX.exists():
        sys.exit(f"missing {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    models = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            print(f"skip template: {name}", file=sys.stderr)
            continue
        m = parse_sheet(wb[name])
        if m is None:
            print(f"skip (no taper): {name}", file=sys.stderr)
            continue
        models.append(m)
        print(f"{name}: {len(m['dimensions'])} pts "
              f"(tip {len(m['provenance']['sections']['tip'])}, "
              f"butt {len(m['provenance']['sections']['butt'])})", file=sys.stderr)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(models, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({len(models)} models)", file=sys.stderr)


if __name__ == "__main__":
    main()
